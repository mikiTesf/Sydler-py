import datetime

from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color

from sydler.utils.populate import Populate
from sydler.utils.role import Role
from sydler.data.member import Member


class ExcelFileGenerator:

    def __init__(self, assignment_queue):
        self.__assignment_queue = assignment_queue
        self.__workbook = Workbook()
        self.active_sheet = self.__workbook.get_active_sheet()
        # program dates
        self.__program_dates = []
        for _assignment in self.__assignment_queue:
            if _assignment.assignment_date not in self.__program_dates:
                self.__program_dates.append(_assignment.assignment_date)
        list.sort(self.__program_dates)
        # assignee lists
        self.__stage_assignees = []
        self.__round_1_left_assignees = []
        self.__round_1_right_assignees = []
        self.__round_2_left_assignees = []
        self.__round_2_right_assignees = []
        self.__second_hall_assignees = []
        # The only member related information the assignment objects come with is member ID.
        # Therefore, there should be a dictionary from which the corresponding member name
        # can be fetched
        self.name_id_pair = dict()
        for _member in Member.select():
            if _member.first_name_is_duplicate:
                self.name_id_pair[_member.ID] = _member.first_name + ' ' + _member.last_name
            self.name_id_pair[_member.id] = _member.first_name

    def sort_assignees(self):
        for _assignment in self.__assignment_queue:
            if _assignment.target_role == Role.STAGE:
                self.__stage_assignees.append(_assignment)
            elif _assignment.target_role == Role.MIC_ROUND_1_LEFT:
                self.__round_1_left_assignees.append(_assignment)
            elif _assignment.target_role == Role.MIC_ROUND_1_RIGHT:
                self.__round_1_right_assignees.append(_assignment)
            elif _assignment.target_role == Role.MIC_ROUND_2_LEFT:
                self.__round_2_left_assignees.append(_assignment)
            elif _assignment.target_role == Role.MIC_ROUND_2_RIGHT:
                self.__round_2_right_assignees.append(_assignment)
            else:
                self.__second_hall_assignees.append(_assignment)
        # sort the assignment queues based on their `assignment_date` attribute
        list.sort(self.__stage_assignees, key=lambda _assignment: _assignment.assignment_date)
        list.sort(self.__round_1_left_assignees, key=lambda _assignment: _assignment.assignment_date)
        list.sort(self.__round_1_right_assignees, key=lambda _assignment: _assignment.assignment_date)
        list.sort(self.__round_2_left_assignees, key=lambda _assignment: _assignment.assignment_date)
        list.sort(self.__round_2_right_assignees, key=lambda _assignment: _assignment.assignment_date)
        list.sort(self.__second_hall_assignees, key=lambda _assignment: _assignment.assignment_date)

    def fill_page_title(self):
        self.active_sheet['A1'].value = 'የአዲስ ሰፈር ጉባኤ\nየድምጽ ክፍል ፕሮግራም'
        self.active_sheet.merge_cells('A1:H1')
        self.format_header_section()

    def format_header_section(self):
        title_font = Font(size=25, bold=True)
        title_alignment = Alignment(horizontal="center", vertical="center")
        self.active_sheet['A1'].font = title_font
        self.active_sheet['A1'].alignment = title_alignment

    def make_excel(self):
        self.sort_assignees()
        self.fill_page_title()
        self.fill_column_titles()
        self.fill_week_spans()
        self.fill_meeting_day_names()
        self.fill_assignees()
        self.complete_generation()

    def fill_column_titles(self):
        week_span_column_title = 'ሳምንት'
        week_span_title_cell = 'A2'
        self.active_sheet[week_span_title_cell] = week_span_column_title

        meeting_day_name_column_title = 'ቀን'
        meeting_day_title_cell = 'B2'
        self.active_sheet[meeting_day_title_cell] = meeting_day_name_column_title

        stage_column_title = 'መድረክ'
        stage_title_cell = 'C2'
        self.active_sheet[stage_title_cell] = stage_column_title

        mic_first_round_title = 'በመጀመሪያ ዙር'
        mic_first_round_title_cell = 'D2'
        self.active_sheet[mic_first_round_title_cell] = mic_first_round_title
        self.active_sheet.merge_cells('D2:E2')

        mic_second_round_title = 'በሁለተኛ ዙር'
        mic_second_round_title_cell = 'F2'
        self.active_sheet[mic_second_round_title_cell] = mic_second_round_title
        self.active_sheet.merge_cells('F2:G2')

        second_hall_title_title = 'ሁለተኛ አዳራሽ'
        second_hall_title_cell = 'H2'
        self.active_sheet[second_hall_title_cell] = second_hall_title_title
        self.format_column_titles()

    def format_column_titles(self):
        title_font = Font(size=12, bold=True)
        alignment = Alignment(horizontal='center', vertical='center')
        fill_color = PatternFill(patternType='solid', fgColor=Color(rgb="BBBBBB"))
        border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

        for column in range(1, 9):
            self.active_sheet.cell(2, column).font = title_font
            self.active_sheet.cell(2, column).alignment = alignment
            self.active_sheet.cell(2, column).fill = fill_color
            self.active_sheet.cell(2, column).border = border

    def fill_week_spans(self):
        row = 3
        for arr_index in range(0, len(self.__program_dates), 2):
            # * Every even index contains a mid-week meeting date
            # * Every odd index contains a weekend meeting date
            # * Therefore to find the beginning of the week (the date on the week's Monday)
            # the day of any mid-week meeting date must be reduced by the number of days
            # it extends from the last Monday
            week_beginning_date = self.__program_dates[arr_index] + datetime.timedelta(
                days=-self.__program_dates[arr_index].weekday())
            week_ending_date = self.__program_dates[arr_index + 1]
            month_on_monday = week_beginning_date.strftime('%b')
            day_on_monday = week_beginning_date.day
            if week_ending_date.month != week_beginning_date.month:
                month_on_weekend = week_ending_date.strftime('%b ')
            else:
                month_on_weekend = ''
            day_on_weekend = week_ending_date.day
            week_span = '{} {} - {}{}'.format(month_on_monday,
                                              day_on_monday,
                                              month_on_weekend,
                                              day_on_weekend)
            begin_cell = 'A' + str(row)
            end_cell = 'A' + str(row + 1)
            self.active_sheet[begin_cell].value = week_span
            self.active_sheet.merge_cells(begin_cell + ":" + end_cell)
            row += 2
        self.format_week_span_cells()

    def format_week_span_cells(self):
        _border = Border(right=Side('thin'), left=Side('thin'), top=Side('thin'))
        _alignment = Alignment(vertical='center', horizontal='center')
        _fill = PatternFill(patternType='solid', fgColor=Color(rgb="BBBBBB"))
        # because `range(...)` doesn't include the upper bound when generating values,
        # the upper bound of the next loop, which increments by 2 on every iteration,
        # must be increased by 2. This increment is important because `row_index` is
        # initially 3 which creates an offset of +2 when compared to the index of the row
        # at which the last stage assignee is found.
        last_row_index = 0
        # due to  the reason above, the row that should be styled in order to fully border
        # the last week-span cell must change from the 'top-left cell' to 'the 'bottom-left cell'
        # (read the documentation for openpyxl for the details of cell styling). That's why
        # the `last_row_index` variable above is important
        for row_index in range(3, len(self.__stage_assignees) + 2, 2):
            cell = 'A' + str(row_index)
            self.active_sheet[cell].fill = _fill
            self.active_sheet[cell].alignment = _alignment
            self.active_sheet[cell].border = _border
            last_row_index = row_index + 1
        _border = Border(bottom=Side('thin'))
        self.active_sheet['A' + str(last_row_index)].border = _border

    def fill_meeting_day_names(self):
        column = 'B'
        row = 3

        for _date in self.__program_dates:
            self.active_sheet[column + str(row)].value = _date.strftime('%A')
            row += 1
        self.format_meeting_day_cells()

    def format_meeting_day_cells(self):
        _alignment = Alignment(horizontal='center', vertical='center')
        _border = Border(bottom=Side('thin'), right=Side('thin'))

        column = 'B'
        row = 3

        for _index in range(len(self.__program_dates)):
            self.active_sheet[column + str(row)].alignment = _alignment
            self.active_sheet[column + str(row)].border = _border
            row += 1

    def fill_assignees(self):
        row_increment = 1

        for role_ in Role.get_roles():
            row = 3

            if role_ == Role.STAGE:
                assignee_queue = self.__stage_assignees
                column = 'C'
            elif role_ == Role.MIC_ROUND_1_LEFT:
                assignee_queue = self.__round_1_left_assignees
                column = 'D'
            elif role_ == Role.MIC_ROUND_1_RIGHT:
                assignee_queue = self.__round_1_right_assignees
                column = 'E'
            elif role_ == Role.MIC_ROUND_2_LEFT:
                assignee_queue = self.__round_2_left_assignees
                column = 'F'
            elif role_ == Role.MIC_ROUND_2_RIGHT:
                assignee_queue = self.__round_2_right_assignees
                column = 'G'
            else:
                assignee_queue = self.__second_hall_assignees
                column = 'H'
                row_increment = 2

            for assignment in assignee_queue:
                assignment_cell = column + str(row)
                self.active_sheet[assignment_cell].value = self.name_id_pair[assignment.assignee_id]
                row += row_increment
        self.format_assignee_cells()

    def format_assignee_cells(self):
        _alignment = Alignment(horizontal='center', vertical='center')
        _border = Border(bottom=Side('thin'), right=Side('thin'))

        for role_ in Role.get_roles():

            if role_ == Role.STAGE:
                column = 'C'
            elif role_ == Role.MIC_ROUND_1_LEFT:
                column = 'D'
            elif role_ == Role.MIC_ROUND_1_RIGHT:
                column = 'E'
            elif role_ == Role.MIC_ROUND_2_LEFT:
                column = 'F'
            elif role_ == Role.MIC_ROUND_2_RIGHT:
                column = 'G'
            else:
                column = 'H'

            for row_index in range(1, len(self.__stage_assignees) + 1):
                assignment_cell = column + str(row_index + 2)
                self.active_sheet[assignment_cell].alignment = _alignment
                self.active_sheet[assignment_cell].border = _border

    def complete_generation(self):
        self.active_sheet.paper_size = 'A4'
        self.__workbook.save("../schedule.xlsx")
