import datetime

import openpyxl.workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Color

from controller.populate import Populate
from controller.role import Role
from data.db_connection import DBConnection


class ExcelFileGenerator:

    def __init__(self, assignment_queue):
        self.__assignment_queue = assignment_queue
        self.__workbook = openpyxl.Workbook()
        self.active_sheet = self.__workbook.get_active_sheet()
        # program dates
        self.__program_dates = []
        for _assignment in self.__assignment_queue:
            if _assignment.assignment_date not in self.__program_dates:
                self.__program_dates.append(_assignment.assignment_date)
        list.sort(self.__program_dates)
        # assignee lists
        self.__stage_assignees = []
        self.__round_1_left_assignees = []
        self.__round_1_right_assignees = []
        self.__round_2_left_assignees = []
        self.__round_2_right_assignees = []
        self.__second_hall_assignees = []
        # The only member related information the assignment objects come with is member ID.
        # Therefore, there should be a dictionary from which the corresponding member name
        # can be fetched
        self.name_id_pair = dict()
        for _member in DBConnection.get_all_members():
            if _member.first_name_is_duplicate:
                self.name_id_pair[_member.ID] = _member.first_name + ' ' + _member.last_name
            self.name_id_pair[_member.ID] = _member.first_name

    def sort_assignees(self):
        for _assignment in self.__assignment_queue:
            if _assignment.target_role == Role.STAGE:
                self.__stage_assignees.append(_assignment)
            elif _assignment.target_role == Role.MIC_ROUND_1_LEFT:
                self.__round_1_left_assignees.append(_assignment)
            elif _assignment.target_role == Role.MIC_ROUND_1_RIGHT:
                self.__round_1_right_assignees.append(_assignment)
            elif _assignment.target_role == Role.MIC_ROUND_2_LEFT:
                self.__round_2_left_assignees.append(_assignment)
            elif _assignment.target_role == Role.MIC_ROUND_2_RIGHT:
                self.__round_2_right_assignees.append(_assignment)
            else:
                self.__second_hall_assignees.append(_assignment)
        # sort the assignment queues based on their `assignment_date` attribute
        list.sort(self.__stage_assignees, key=lambda _assignment: _assignment.assignment_date)
        list.sort(self.__round_1_left_assignees, key=lambda _assignment: _assignment.assignment_date)
        list.sort(self.__round_1_right_assignees, key=lambda _assignment: _assignment.assignment_date)
        list.sort(self.__round_2_left_assignees, key=lambda _assignment: _assignment.assignment_date)
        list.sort(self.__round_2_right_assignees, key=lambda _assignment: _assignment.assignment_date)
        list.sort(self.__second_hall_assignees, key=lambda _assignment: _assignment.assignment_date)

    def fill_header_info(self):
        a1 = self.active_sheet['A1']
        a1.value = 'የአዲስ ሰፈር ጉባኤ\nየድምጽ ክፍል ፕሮግራም'
        self.active_sheet.merge_cells('A1:H1')
        title_font = Font(size=25, bold=True)
        title_alignment = Alignment(horizontal="center", vertical="center")
        a1.font = title_font
        a1.alignment = title_alignment

    def make_excel(self):
        self.sort_assignees()
        self.fill_header_info()
        self.fill_column_titles()
        self.fill_week_spans()
        self.fill_meeting_day_names()
        self.fill_assignees()
        self.complete_generation()

    def fill_column_titles(self):
        title_font = Font(size=12, bold=True)
        alignment = Alignment(horizontal='center', vertical='center')
        fill_color = PatternFill(patternType='solid', fgColor=Color(rgb="BBBBBB"))
        border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

        week_span_column_title = 'ሳምንት'
        week_span_title_cell = 'A2'
        self.active_sheet[week_span_title_cell] = week_span_column_title

        meeting_day_name_column_title = 'ቀን'
        meeting_day_title_cell = 'B2'
        self.active_sheet[meeting_day_title_cell] = meeting_day_name_column_title

        stage_column_title = 'መድረክ'
        stage_title_cell = 'C2'
        self.active_sheet[stage_title_cell] = stage_column_title

        mic_first_round_title = 'በመጀመሪያ ዙር'
        mic_first_round_title_cell = 'D2'
        self.active_sheet[mic_first_round_title_cell] = mic_first_round_title
        self.active_sheet.merge_cells('D2:E2')

        mic_second_round_title = 'በሁለተኛ ዙር'
        mic_second_round_title_cell = 'F2'
        self.active_sheet[mic_second_round_title_cell] = mic_second_round_title
        self.active_sheet.merge_cells('F2:G2')

        second_hall_title_title = 'ሁለተኛ አዳራሽ'
        second_hall_title_cell = 'H2'
        self.active_sheet[second_hall_title_cell] = second_hall_title_title

        for column in range(1, 9):
            self.active_sheet.cell(2, column).font = title_font
            self.active_sheet.cell(2, column).alignment = alignment
            self.active_sheet.cell(2, column).fill = fill_color
            self.active_sheet.cell(2, column).border = border

    def fill_week_spans(self):
        _border = Border(right=Side('thin'), left=Side('thin'), bottom=Side('thin'))
        _alignment = Alignment(vertical='center', horizontal='center')
        _fill = PatternFill(patternType='solid', fgColor=Color(rgb="BBBBBB"))

        row = 3
        for _index in range(0, len(self.__program_dates) - 1, 2):
            # * Every even index contains a mid-week meeting date
            # * Every odd index contains a weekend meeting date
            # * Therefore to find the beginning of the week (the date on the week's Monday)
            # the day of any mid-week meeting date must be reduced by the number of days
            # it extends from the last Monday
            week_beginning_date = self.__program_dates[_index] + datetime.timedelta(days=-self.__program_dates[_index].weekday())
            week_ending_date = self.__program_dates[_index + 1]
            month_on_monday = week_beginning_date.strftime('%b')
            day_on_monday = week_beginning_date.day
            if week_ending_date.month != week_beginning_date.month:
                month_on_weekend = week_ending_date.strftime('%b ')
            else:
                month_on_weekend = ''
            day_on_weekend = week_ending_date.day
            week_span = '{} {} - {}{}'.format(month_on_monday,
                                              day_on_monday,
                                              month_on_weekend,
                                              day_on_weekend)
            cell = 'A' + str(row)
            self.active_sheet[cell].value = week_span
            self.active_sheet[cell].border = _border
            self.active_sheet[cell].alignment = _alignment
            self.active_sheet[cell].fill = _fill
            self.active_sheet.merge_cells(cell + ":A" + str(row + 1))
            row += 2

    def fill_meeting_day_names(self):
        column = 'B'
        row = 3
        _alignment = Alignment(horizontal='center', vertical='center')
        _border = Border(bottom=Side('thin'), right=Side('thin'))

        for _date in self.__program_dates:
            self.active_sheet[column + str(row)].value = _date.strftime('%A')
            self.active_sheet[column + str(row)].alignment = _alignment
            self.active_sheet[column + str(row)].border = _border
            row += 1

    def fill_assignees(self):
        row_increment = 1
        _alignment = Alignment(horizontal='center', vertical='center')
        _border = Border(bottom=Side('thin'), right=Side('thin'))

        for _role in Role.get_roles():
            row = 3

            if _role == Role.STAGE:
                assignee_queue = self.__stage_assignees
                column = 'C'
            elif _role == Role.MIC_ROUND_1_LEFT:
                assignee_queue = self.__round_1_left_assignees
                column = 'D'
            elif _role == Role.MIC_ROUND_1_RIGHT:
                assignee_queue = self.__round_1_right_assignees
                column = 'E'
            elif _role == Role.MIC_ROUND_2_LEFT:
                assignee_queue = self.__round_2_left_assignees
                column = 'F'
            elif _role == Role.MIC_ROUND_2_RIGHT:
                assignee_queue = self.__round_2_right_assignees
                column = 'G'
            else:
                assignee_queue = self.__second_hall_assignees
                column = 'H'
                row_increment = 2

            for _assignment in assignee_queue:
                assignment_cell = column + str(row)
                self.active_sheet[assignment_cell].value = self.name_id_pair[_assignment.assignee_id]
                self.active_sheet[assignment_cell].alignment = _alignment
                self.active_sheet[assignment_cell].border = _border
                row += row_increment

    def complete_generation(self):
        self.active_sheet.paper_size = 'A4'
        self.__workbook.save("../schedule.xlsx")


# `day` below is a Tuesday
day = datetime.date(2019, 4, 9)
dates = [day]

days_to_add = 5
for index in range(1, 32):
    day = day + datetime.timedelta(days=days_to_add)
    dates.append(day)
    days_to_add = 2 if days_to_add == 5 else 5

populate = Populate(dates)
file_generator = ExcelFileGenerator(populate.get_assignments())
file_generator.make_excel()
